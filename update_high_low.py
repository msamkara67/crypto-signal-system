import os
from config import working_dir_name

desktop = os.path.expanduser("~/Desktop")
base_dir = os.path.join(desktop, working_dir_name)
from openpyxl import load_workbook
from datetime import datetime, timedelta
import yfinance as yf
import os

# === Dosya yolları ===
from config import template_path, data_path

sheet_name = "Daily Update"

# === Excel şablonunu aç ===
wb_template = load_workbook(template_path)
ws_template = wb_template[sheet_name]

# === B1 hücresinden tarihi al ===
date_cell = ws_template["B1"].value
if not isinstance(date_cell, datetime):
    date_cell = datetime.strptime(str(date_cell), "%Y-%m-%d")
target_date = date_cell.date()

start_str = target_date.strftime("%Y-%m-%d")
end_str = (target_date + timedelta(days=1)).strftime("%Y-%m-%d")

print(f"🎯 Tarih: {start_str} için veri çekiliyor...\n")

# === Coin listesini oku ===
coins = []
row_index = 2
while True:
    coin = ws_template.cell(row=row_index, column=3).value
    if coin is None or str(coin).strip() == "":
        break
    coins.append((row_index, coin))
    row_index += 1

# === High / Low verilerini çek ===
for row_index, coin in coins:
    ticker_name = f"{coin}-USD"
    try:
        ticker = yf.Ticker(ticker_name)
        hist = ticker.history(start=start_str, end=end_str)

        if not hist.empty:
            high_val = "{:.8f}".format(hist["High"].iloc[0])
            low_val = "{:.8f}".format(hist["Low"].iloc[0])
            print(f"✅ {ticker_name} → High: {high_val} | Low: {low_val}")
        else:
            high_val = "NaN"
            low_val = "NaN"
            print(f"⚠️ {ticker_name} → Veri bulunamadı. Ticker geçersiz olabilir veya bu tarihte veri yok.")

    except Exception as e:
        high_val = "HATA"
        low_val = "HATA"
        print(f"❌ {ticker_name} → Hata: {e}")

    ws_template.cell(row=row_index, column=4).value = high_val  # D sütunu = High
    ws_template.cell(row=row_index, column=5).value = low_val   # E sütunu = Low

wb_template.save(template_path)
print("\n✅ Template dosyası (High/Low) güncellendi.")

# === 180 günlük veriyi aç ===
wb_data = load_workbook(data_path)
ws_high = wb_data["Daily Update (High)"]
ws_low = wb_data["Daily Update (Low)"]

# === Tarihe karşılık gelen sütunu bul ===
def find_column_by_date(ws, date_val):
    for col in range(4, ws.max_column + 1):  # D sütunundan başlar
        val = ws.cell(row=1, column=col).value
        if isinstance(val, datetime) and val.date() == date_val:
            return col
    return None

target_col = find_column_by_date(ws_high, target_date)
if not target_col:
    raise ValueError(f"🛑 Hedef tarih bulunamadı: {target_date}")

# === Verileri ana dosyaya yaz ===
for row_index, coin in coins:
    high_val = ws_template.cell(row=row_index, column=4).value
    low_val = ws_template.cell(row=row_index, column=5).value

    for r in range(2, ws_high.max_row + 1):
        if ws_high.cell(r, 3).value == coin:
            ws_high.cell(r, target_col).value = high_val
            ws_low.cell(r, target_col).value = low_val
            break

wb_data.save(data_path)
print("✅ High/Low verileri ana 180 günlük dosyaya başarıyla aktarıldı.")
