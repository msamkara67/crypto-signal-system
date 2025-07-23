import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
from openpyxl import load_workbook
from datetime import datetime
import sys

# Excel yolu
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
sheet_name = "Daily Update"

def select_date_and_write():
    def on_select():
        date_str = cal.get_date()
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

            # Excel dosyasına yaz
            wb = load_workbook(template_path)
            ws = wb[sheet_name]
            ws["B1"].value = date_obj
            wb.save(template_path)
            wb.close()

            print(f"{date_obj}")  # stdout'a yaz, ana kod bunu okuyacak
            root.destroy()
        except Exception as e:
            messagebox.showerror("Hata", f"Tarih işlenemedi:\n{e}")

    root = tk.Tk()
    root.title("Tarih Seç")
    root.geometry("300x250")
    root.resizable(False, False)

    cal = Calendar(root, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack(padx=10, pady=10)

    btn = tk.Button(root, text="Tamam", command=on_select)
    btn.pack(pady=10)

    root.mainloop()

# Ana fonksiyonu çalıştır
if __name__ == "__main__":
    select_date_and_write()




