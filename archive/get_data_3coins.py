from openpyxl import Workbook
from datetime import datetime, timedelta
import yfinance as yf

# === Coin listesi ve özel tickerlar ===
coins = ["SUI1", "PYTH24625"]
custom_tickers = {
    "SUI1": "SUI1-USD",
    "PYTH24625": "PYTH24625-USD"
}

days_back = 180
end_date = datetime.today().date()
start_date = end_date - timedelta(days=days_back)

# === Excel oluştur (Close + Volume ayrı sayfalarda)
wb = Workbook()

# Sheet 1: Close
ws_close = wb.active
ws_close.title = "Daily Update (Close)"
ws_close.cell(row=1, column=1).value = "Coins"

# Sheet 2: Volume
ws_volume = wb.create_sheet("Daily Update (Volume)")
ws_volume.cell(row=1, column=1).value = "Coins"

# === Tarih başlıkları
for i in range(days_back):
    date = start_date + timedelta(days=i)
    ws_close.cell(row=1, column=2 + i).value = date
    ws_volume.cell(row=1, column=2 + i).value = date

# === Veri çek ve yaz
for idx, coin in enumerate(coins):
    ticker_symbol = custom_tickers[coin]
    print(f"🔄 {coin} veri çekiliyor → {ticker_symbol}")
    
    try:
        ticker = yf.Ticker(ticker_symbol)
        hist = ticker.history(start=start_date, end=end_date)

        row = idx + 2
        ws_close.cell(row=row, column=1).value = coin
        ws_volume.cell(row=row, column=1).value = coin

        for i in range(days_back):
            date = start_date + timedelta(days=i)
            str_date = date.strftime("%Y-%m-%d")

            close_val = hist["Close"].get(str_date)
            vol_val = hist["Volume"].get(str_date)

            if close_val is not None:
                ws_close.cell(row=row, column=2 + i).value = float("{:.8f}".format(close_val))
            if vol_val is not None:
                ws_volume.cell(row=row, column=2 + i).value = int(vol_val)

        print(f"✅ {coin} başarıyla işlendi.")

    except Exception as e:
        print(f"❌ {coin} veri çekilemedi: {e}")

# === Dosyayı kaydet
save_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\historical_SUI_PYTH_data.xlsx"
wb.save(save_path)
print(f"\n📁 Veriler kaydedildi: {save_path}")



