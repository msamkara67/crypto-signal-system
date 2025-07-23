from openpyxl import load_workbook
from datetime import datetime, timedelta
import yfinance as yf
import os
import shutil
# === Date Pick Up ===

import subprocess

# Pop-up'ı çalıştır ve girilen tarihi al
result = subprocess.run(['python', 'date_picker.py'], capture_output=True, text=True)

# Pop-up'tan gelen tarihi al
target_date = result.stdout.strip()



if not target_date:
    raise ValueError("🛑 Geçerli bir tarih girilmedi.")

print(f"🎯 Veri çekilecek tarih: {target_date}")

# === Dosya Yolları ===
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
data_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_180days_top100.xlsx"
backup_path = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto\coin_data_180days_top100.xlsx"
sheet_name = "Daily Update"

print("=== Günlük veri çekiliyor ve aktarılıyor ===")
wb_template = load_workbook(template_path)
ws_template = wb_template[sheet_name]

# === Tarihi Al ===
date_cell = ws_template.cell(row=1, column=2).value
if isinstance(date_cell, datetime):
    target_date = date_cell.date()
else:
    target_date = datetime.strptime(str(date_cell), "%Y-%m-%d").date()

start_str = target_date.strftime("%Y-%m-%d")
end_str = (target_date + timedelta(days=1)).strftime("%Y-%m-%d")

# === Coin Listesi ===
coins = []
row_index = 2
while True:
    coin = ws_template.cell(row=row_index, column=3).value
    if coin is None or str(coin).strip() == "":
        break
    coins.append((row_index, coin))
    row_index += 1

# === Veri Çek ve Yaz ===
for row_index, coin in coins:
    try:
        ticker = yf.Ticker(f"{coin}-USD")
        hist = ticker.history(start=start_str, end=end_str)

        close_val = "{:.8f}".format(hist["Close"].iloc[0]) if not hist["Close"].isna().all() else "NaN"
        vol_val = "{:.2f}".format(hist["Volume"].iloc[0]) if not hist["Volume"].isna().all() else "NaN"

        print(f"{coin} → Close: {close_val} | Volume: {vol_val}")
    except Exception as e:
        close_val = "HATA"
        vol_val = "HATA"
        print(f"❌ {coin} veri çekilemedi: {e}")

    ws_template.cell(row=row_index, column=4).value = close_val
    ws_template.cell(row=row_index, column=5).value = vol_val

wb_template.save(template_path)
print("✔ Template dosyası güncellendi.")

# === 180 Günlük Dosyayı Güncelle ===
wb_data = load_workbook(data_path)
ws_close = wb_data["Daily Update (Close)"]
ws_volume = wb_data["Daily Update (Volume)"]

def find_column_by_date(ws, date_val):
    for col in range(4, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, datetime) and val.date() == date_val:
            return col
    return None

target_col = find_column_by_date(ws_close, target_date)
if not target_col:
    raise ValueError(f"🛑 Hedef tarih bulunamadı: {target_date}")

for row_index, coin in coins:
    close_val = ws_template.cell(row=row_index, column=4).value
    vol_val = ws_template.cell(row=row_index, column=5).value

    for r in range(2, ws_close.max_row + 1):
        if ws_close.cell(row=r, column=3).value == coin:
            ws_close.cell(row=r, column=target_col).value = close_val
            ws_volume.cell(row=r, column=target_col).value = vol_val
            break

wb_data.save(data_path)
print("✅ Veriler ana 180 günlük dosyaya aktarıldı.")

# --- Arşivleme ---
subprocess.run(['python', 'archive_backup.py', target_date.strftime('%Y-%m-%d')])
# --- Arşivleme ---





