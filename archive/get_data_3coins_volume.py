from openpyxl import Workbook
from datetime import datetime, timedelta
import yfinance as yf

# === Coin ve Ticker Ayarları ===
coins = ["LDO", "PEPE24478"]
custom_tickers = {
    "PEPE24478": "PEPE24478-USD",
    "LDO": "LDO-USD"
}

days_back = 180
end_date = datetime.today().date()
start_date = end_date - timedelta(days=days_back)

# === Excel dosyası oluştur
wb = Workbook()
ws = wb.active
ws.title = "Daily Update (Volume)"

# === Başlık satırı (tarihler)
ws.cell(row=1, column=1).value = "Coins"
for i in range(days_back):
    date = start_date + timedelta(days=i)
    ws.cell(row=1, column=2 + i).value = date

# === Coin bazlı Volume verisi çek ve yaz
for idx, coin in enumerate(coins):
    ticker_symbol = custom_tickers[coin]
    print(f"🔄 {coin} için volume verisi çekiliyor → {ticker_symbol}")

    try:
        ticker = yf.Ticker(ticker_symbol)
        hist = ticker.history(start=start_date, end=end_date)

        row = idx + 2
        ws.cell(row=row, column=1).value = coin

        for i in range(days_back):
            date = start_date + timedelta(days=i)
            volume = hist["Volume"].get(date.strftime("%Y-%m-%d"), None)
            if volume is not None:
                ws.cell(row=row, column=2 + i).value = int(volume)

        print(f"✅ {coin} volume verisi başarıyla işlendi.")

    except Exception as e:
        print(f"❌ {coin} için volume verisi çekilemedi: {e}")

# === Dosyayı kaydet
save_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\historical_volume_data_LDO_PEPE24478.xlsx"
wb.save(save_path)
print(f"\n📁 Volume verileri kaydedildi: {save_path}")



