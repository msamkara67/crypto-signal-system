# coin_updater_debug_FULL.py

from openpyxl import load_workbook
from datetime import datetime
import os
import shutil
from tkinter import Tk, simpledialog

print("[RUN] Sistem baslatiliyor...")

# [0] Tarih seçimi popup
root = Tk()
root.withdraw()
selected_date_str = simpledialog.askstring("Tarih Secimi", "Tarih girin (YYYY-MM-DD):")
if not selected_date_str:
    print("[ERR] Tarih girilmedi, cikiliyor.")
    exit()
selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
print(f"[1] Secilen tarih: {selected_date}")

# [1] Dosya yolları
SOURCE_FILE = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
TARGET_FILE = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto\coin_data_180days_top100.xlsx"

# [2] Dosyaları yükle
source_wb = load_workbook(SOURCE_FILE)
target_wb = load_workbook(TARGET_FILE)

source_ws = source_wb.active
close_ws = target_wb["Daily Update (Close)"]
volume_ws = target_wb["Daily Update (Volume)"]

# [3] Tarih sütununu bul
header_row = 1
found_col = None
for col in range(3, source_ws.max_column + 1):
    val = source_ws.cell(row=header_row, column=col).value
    if isinstance(val, datetime):
        val = val.date()
    elif isinstance(val, str):
        try:
            val = datetime.strptime(val, "%Y-%m-%d").date()
        except:
            continue
    if val == selected_date:
        found_col = col
        break

if not found_col:
    print("[ERR] Kaynak dosyada tarih bulunamadi.")
    exit()

print(f"[2] Tarih sütunu bulundu: {found_col}")

# [4] Coin listesini oku
coins = []
for row in range(2, source_ws.max_row + 1):
    coin = source_ws.cell(row=row, column=3).value
    if coin:
        coins.append((row, coin))

# [5] Hedef dosyada ilgili tarihi bul
col_target = None
for col in range(4, close_ws.max_column + 1):
    val = close_ws.cell(row=1, column=col).value
    if isinstance(val, datetime):
        val = val.date()
    elif isinstance(val, str):
        try:
            val = datetime.strptime(val, "%Y-%m-%d").date()
        except:
            continue
    if val == selected_date:
        col_target = col
        break

if not col_target:
    print("[ERR] Hedef dosyada tarih sütunu bulunamadi.")
    exit()

print(f"[3] Hedef dosyada tarih sütunu bulundu: {col_target}")

# [6] Close ve Volume verilerini aktar
update_count = 0
for row_source, coin in coins:
    close_val = source_ws.cell(row=row_source, column=4).value
    volume_val = source_ws.cell(row=row_source, column=5).value

    # Coin'i hedef dosyada bul
    target_row = None
    for row in range(2, close_ws.max_row + 1):
        coin_target = close_ws.cell(row=row, column=2).value
        if coin_target == coin:
            target_row = row
            break

    if target_row:
        close_ws.cell(row=target_row, column=col_target).value = close_val
        volume_ws.cell(row=target_row, column=col_target).value = volume_val
        update_count += 1

print(f"[4] {update_count} coin güncellendi.")

# [7] Kayıt
target_wb.save(TARGET_FILE)
print("[5] ✔ Dosya kaydedildi.")

