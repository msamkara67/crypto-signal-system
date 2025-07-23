from openpyxl import load_workbook
from datetime import datetime

# Excel dosyasının yolu
excel_path = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto\coin_data_180days_top100.xlsx"

# Hedef tarih
target_date_str = "2025-07-15"
target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()

# Dosya aç
wb = load_workbook(excel_path)
coins_ws = wb["Coins"]
rsi_ws = wb["RSI"]

# Tarih satırı 2. satır
target_col = None
for col in range(5, coins_ws.max_column + 1):
    val = coins_ws.cell(row=2, column=col).value
    if isinstance(val, datetime) and val.date() == target_date:
        target_col = col
        break

if not target_col:
    print("Tarih sütunu bulunamadı.")
    wb.close()
    exit()

print(f"[1] İşleme alınan tarih: {target_date_str}")
print(f"[2] Coins sayfasında tarih sütunu bulundu: {target_col}")

signal_count = 0
for row in range(3, coins_ws.max_row + 1):
    rsi_vals = []
    vol_vals = []

    for offset in range(6):
        rsi_val = coins_ws.cell(row=row, column=target_col - offset).value
        vol_val = coins_ws.cell(row=row, column=target_col - offset + 1).value
        rsi_vals.append(rsi_val if isinstance(rsi_val, (int, float)) else None)
        vol_vals.append(vol_val if isinstance(vol_val, (int, float)) else None)

    rsi_today = rsi_vals[0]
    vol_today = vol_vals[0]
    if rsi_today is None or vol_today is None:
        continue

    ### R: RSI <30 → AL, >70 → SAT
    r_signal = "AL" if rsi_today < 30 else "SAT" if rsi_today > 70 else ""

    ### S: 30 < RSI < 70 ve son 3 günde dramatik RSI değişimi
    s_signal = ""
    if 30 <= rsi_today <= 70 and all(r is not None for r in rsi_vals[:4]):
        diffs = [abs(rsi_vals[i] - rsi_vals[i+1]) for i in range(3)]
        if max(diffs) >= 20:
            s_signal = "FAKAT"

    ### T: Günlük hacim > 1.5 * 5 günlük ortalama hacim
    t_signal = ""
    valid_vols = [v for v in vol_vals[1:6] if v is not None]
    if valid_vols:
        avg_vol = sum(valid_vols) / len(valid_vols)
        if vol_today >= 1.5 * avg_vol:
            t_signal = "!!!"

    if r_signal or s_signal or t_signal:
        signal_count += 1

    rsi_ws.cell(row=row, column=18).value = r_signal
    rsi_ws.cell(row=row, column=19).value = s_signal
    rsi_ws.cell(row=row, column=20).value = t_signal

print(f"[3] {signal_count} sinyal üretildi.")
print("[4] Sinyaller kaydedildi.")

wb.save(excel_path)
wb.close()
print("[DONE] Tüm işlem başarıyla tamamlandı.")


