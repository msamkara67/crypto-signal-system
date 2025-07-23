import tkinter as tk
from tkinter import simpledialog
from openpyxl import load_workbook
from datetime import datetime

# === Excel yolu ===
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
sheet_name = "Daily Update"

# === Pop-up ile tarih al ===
def get_target_date():
    root = tk.Tk()
    root.withdraw()
    date_input = simpledialog.askstring("Tarih Seç", "Tarihi girin (YYYY-MM-DD):")
    if date_input:
        try:
            date_obj = datetime.strptime(date_input, "%Y-%m-%d")
            return date_obj
        except ValueError:
            print("⚠️ Hatalı format: YYYY-MM-DD olmalı.")
            return None
    return None

# === Çalıştır ===
target_date = get_target_date()
if target_date:
    # Excel dosyasına yaz
    wb = load_workbook(template_path)
    ws = wb[sheet_name]
    ws["B1"] = target_date  # B1'e datetime objesi olarak yaz
    wb.save(template_path)
    wb.close()

    print(target_date.strftime("%Y-%m-%d"))  # stdout → ana kod bu değeri alıyor
else:
    print("")  # Hatalı girişte boş değer dön





