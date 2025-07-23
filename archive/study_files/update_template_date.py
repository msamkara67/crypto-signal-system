from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog, messagebox

# Excel yolu
excel_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"

# GUI başlat
root = tk.Tk()
root.withdraw()  # Pencereyi gizle

# Tarih bilgisi al
input_date = simpledialog.askstring("Tarih Girişi", "Lütfen veri çekmek istediğiniz tarihi girin (yyyy-aa-gg):")

# Cancel tuşuna basıldıysa hiçbir şey yapmadan çık
if input_date is None:
    print("❌ Giriş iptal edildi. Çıkılıyor...")
    exit(0)

# Format kontrolü
try:
    target_date = datetime.strptime(input_date, "%Y-%m-%d").date()
except ValueError:
    messagebox.showerror("Hata", "❌ Hatalı tarih formatı. Örnek: 2025-07-17")
    exit(1)

# Excel dosyasını güncelle
wb = load_workbook(excel_path)
ws = wb["Daily Update"]
ws.cell(row=1, column=2).value = target_date
wb.save(excel_path)

print(f"✅ Tarih başarıyla B1 hücresine yazıldı: {target_date}")
