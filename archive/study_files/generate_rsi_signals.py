# generate_rsi_signals.py

from openpyxl import load_workbook
from datetime import datetime, timedelta
import sys
import os

print("[SIGNAL] Sinyalizasyon başlatılıyor...")

# Ayarlar
MAIN_PATH = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto"
TEMPLATE_PATH = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_180days_top100.xlsx"
TARGET_FILE = os.path.join(MAIN_PATH, "coin_data_180days_top100.xlsx")
DATE_FORMAT = "%Y-%m-%d"

# Eğer hedef dosya yoksa kopyala
if not os.path.exists(TARGET_FILE):
    import shutil
    shutil.copy(TEMPLATE_PATH, TARGET_FILE)
    print("[COPY] coin_data_180days_top100.xlsx dosyası coin_Updater'dan kopyalandı.")

# Tarihi komut satırından al
if len(sys.argv) < 2:
    print("[ERR] Tarih argümanı eksik!")
    sys.exit(1)

input_date_str = sys.argv[1]
input_date = datetime.strptime(input_date_str, DATE_FORMAT).date()

# Excel dosyasını aç
wb = load_workbook(TARGET_FILE)
coins_ws = wb["Coins"]
rsi_ws = wb["RSI"]

# Tarih sütununu bul (2. satır)
def find_column_by_date(ws, sheet_name):
    for col in range(5, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col).value
        if isinstance(cell_value, datetime):
            if cell_value.date() == input_date:
                return col
        elif isinstance(cell_value, str):
            try:
                if datetime.strptime(cell_value, DATE_FORMAT).date() == input_date:
                    return col
            except:
                continue
    print(f"[ERR] {sheet_name} sayfasında tarih bulunamadı: {input_date_str}")
    return None

col_index = find_column_by_date(coins_ws, "Coins")
if not col_index:
    sys.exit(1)
print(f"[2] Coins sayfasında tarih sütunu bulundu: {col_index}")

# RSI14 ve Volume için referansları bul
coin_names = []
rsi_rows = {}
volume_rows = {}
for row in range(3, coins_ws.max_row + 1):
    param = coins_ws.cell(row=row, column=4).value
    coin = coins_ws.cell(row=row, column=3).value
    if not coin:
        continue
    if param == "RSI14":
        coin_names.append(coin)
        rsi_rows[coin] = row
    elif param == "Volume":
        volume_rows[coin] = row

# RSI sayfasında coinleri bul
rsi_name_to_row = {}
for row in range(3, rsi_ws.max_row + 1):
    name = rsi_ws.cell(row=row, column=2).value
    if name:
        rsi_name_to_row[name] = row

# Sinyalleri üret ve yaz
signal_count = 0
for coin in coin_names:
    if coin not in rsi_name_to_row:
        continue

    rsi_row = rsi_rows.get(coin)
    volume_row = volume_rows.get(coin)
    rsi_today = coins_ws.cell(row=rsi_row, column=col_index).value

    rsi_cell_range = [coins_ws.cell(row=rsi_row, column=col_index - i).value for i in range(3)]
    volume_today = coins_ws.cell(row=volume_row, column=col_index).value
    volume_past = [coins_ws.cell(row=volume_row, column=col_index - i).value for i in range(1, 6)]

    # Boş veri varsa atla
    if None in rsi_cell_range or None in volume_past or volume_today is None:
        continue

    row_idx = rsi_name_to_row[coin]
    r_signal, s_signal, t_signal = "", "", ""

    # R sütunu: AL/SAT sinyali
    if rsi_today < 30:
        r_signal = "AL"
    elif rsi_today > 70:
        r_signal = "SAT"

    # S sütunu: FAKAT sinyali (3 gün içinde RSI değişimi ±20)
    delta_rsi = max(rsi_cell_range) - min(rsi_cell_range)
    if 30 < rsi_today < 70 and delta_rsi >= 20:
        s_signal = "FAKAT"

    # T sütunu: Hacim çarpanı
    avg_volume = sum(volume_past) / len(volume_past)
    if volume_today >= 1.5 * avg_volume:
        t_signal = "!!!"

    if r_signal or s_signal or t_signal:
        rsi_ws.cell(row=row_idx, column=18).value = r_signal
        rsi_ws.cell(row=row_idx, column=19).value = s_signal
        rsi_ws.cell(row=row_idx, column=20).value = t_signal
        signal_count += 1

print(f"[3] {signal_count} sinyal üretildi.")
print("[4] ✔ Sinyaller kaydedildi.")
wb.save(TARGET_FILE)

