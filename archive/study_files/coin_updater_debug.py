from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta
import yfinance as yf
import os

# Dosya yolları
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
data_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_180days_top100.xlsx"
sheet_name = "Daily Update"

print("=== Günlük veri çekiliyor ve aktarılıyor ===")
wb_template = load_workbook(template_path)
ws_template = wb_template[sheet_name]

# Tarih al
date_cell = ws_template.cell(row=1, column=2).value
if isinstance(date_cell, datetime):
    target_date = date_cell.date()
else:
    target_date = datetime.strptime(str(date_cell), "%Y-%m-%d").date()

start_str = target_date.strftime("%Y-%m-%d")
end_str = (target_date + timedelta(days=1)).strftime("%Y-%m-%d")

# Coin listesi
coins = []
row_index = 2
while True:
    coin = ws_template.cell(row=row_index, column=3).value
    if coin is None or str(coin).strip() == "":
        break
    coins.append((row_index, coin))
    row_index += 1

# Verileri çek ve yaz
for row_index, coin in coins:
    try:
        ticker = yf.Ticker(f"{coin}-USD")
        hist = ticker.history(start=start_str, end=end_str)

        close_val = round(hist["Close"].iloc[0], 2) if not hist["Close"].isna().all() else "NaN"
        vol_val = round(hist["Volume"].iloc[0], 2) if not hist["Volume"].isna().all() else "NaN"

        print(f"{coin} → Close: {close_val} | Volume: {vol_val}")
    except Exception as e:
        close_val = "HATA"
        vol_val = "HATA"
        print(f"❌ {coin} veri çekilemedi: {e}")

    ws_template.cell(row=row_index, column=4).value = close_val
    ws_template.cell(row=row_index, column=5).value = vol_val

wb_template.save(template_path)
print("✔ Template dosyası güncellendi.")

# 180 günlük dosyaya aktarım
wb_data = load_workbook(data_path)
ws_close = wb_data["Daily Update (Close)"]
ws_volume = wb_data["Daily Update (Volume)"]

# 🔍 Tarih kolonunu her hücre türüne göre tarayan fonksiyon
def find_column_by_date(ws, date_val):
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        # 1. datetime nesnesi mi?
        if isinstance(cell_val, datetime) and cell_val.date() == date_val:
            return col
        # 2. string mi?
        elif isinstance(cell_val, str):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d"):
                try:
                    parsed = datetime.strptime(cell_val.strip(), fmt).date()
                    if parsed == date_val:
                        return col
                except:
                    continue
        # 3. excel float mı?
        elif isinstance(cell_val, (float, int)):
            try:
                parsed = from_excel(cell_val).date()
                if parsed == date_val:
                    return col
            except:
                continue
    return None

#test!
# print("🧪 1. Satırdaki tarih hücreleri:")
# for col in range(1, ws_close.max_column + 1):
#    val = ws_close.cell(row=1, column=col).value
#    print(f"Col {col}: {val} ({type(val)})")

#test!

# Hedef sütunu bul
target_col = find_column_by_date(ws_close, target_date)
if not target_col:
    raise ValueError(f"🛑 Hedef tarih bulunamadı: {target_date}")

# Verileri kopyala
for row_index, coin in coins:
    close_val = ws_template.cell(row=row_index, column=4).value
    vol_val = ws_template.cell(row=row_index, column=5).value

    for r in range(2, ws_close.max_row + 1):
        if ws_close.cell(row=r, column=3).value == coin:
            ws_close.cell(row=r, column=target_col).value = close_val
            ws_volume.cell(row=r, column=target_col).value = vol_val
            break

wb_data.save(data_path)
print("✅ Veriler 180 günlük dosyaya aktarıldı.")
os.startfile(data_path)







