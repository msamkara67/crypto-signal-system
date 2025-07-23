
import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog
import subprocess
import os
from datetime import datetime
from shutil import copyfile
from openpyxl import load_workbook

# === TARİH SEÇİMİ ===
root = tk.Tk()
root.withdraw()
date_input = simpledialog.askstring("Tarih Seç", "İşlenecek tarihi girin (YYYY-MM-DD):")
if not date_input:
    messagebox.showerror("Hata", "Tarih girilmedi.")
    exit()

try:
    target_date = datetime.strptime(date_input, "%Y-%m-%d").date()
except ValueError:
    messagebox.showerror("Hata", "Geçersiz tarih formatı.")
    exit()

print(f"[RUN] Sistem başlatılıyor...")
print(f"[0] Tarih seçimi popup'ı başlatıldı.")
print(f"[1] Seçilen tarih: {target_date}")

# === DOSYA YOLLARI ===
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
main_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_180days_top100.xlsx"
kripto_path = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto\coin_data_180days_top100.xlsx"

# === VERİYİ TEMPLATE'DEN ANA DOSYAYA YAZ ===
try:
    wb_template = load_workbook(template_path)
    ws_template = wb_template.active
    wb_main = load_workbook(main_path)
    close_ws = wb_main["Daily Update (Close)"]
    volume_ws = wb_main["Daily Update (Volume)"]

    header_row = 1
    target_col = None
    for col in range(1, close_ws.max_column + 1):
        cell_value = close_ws.cell(row=header_row, column=col).value
        if isinstance(cell_value, datetime):
            cell_value = cell_value.date()
        elif isinstance(cell_value, str):
            try:
                cell_value = datetime.strptime(cell_value, "%Y-%m-%d").date()
            except ValueError:
                continue
        if cell_value == target_date:
            target_col = col
            break

    if target_col is None:
        print(f"[ERR] {target_date} sütunu bulunamadı.")
        exit(1)

    print(f"[2] {target_date} sütunu bulundu → Sütun: {target_col}")

    coins_loaded = 0
    for row in range(2, ws_template.max_row + 1):
        coin = ws_template.cell(row=row, column=3).value
        close = ws_template.cell(row=row, column=4).value
        volume = ws_template.cell(row=row, column=5).value
        if coin is None:
            continue
        close_ws.cell(row=row, column=target_col).value = close
        volume_ws.cell(row=row, column=target_col).value = volume
        coins_loaded += 1

    print(f"[3] {coins_loaded} coin yüklendi. Yazılıyor...")
    wb_main.save(main_path)
    print(f"[4] ✅ Yazma tamamlandı.")

except Exception as e:
    print(f"[ERR] Veri yazılırken hata: {e}")
    exit(1)

# === DOSYAYI KOPYALA ===
try:
    copyfile(main_path, kripto_path)
    print(f"[5] Dosya kopyalandı → {kripto_path}")
except Exception as e:
    print(f"[ERR] Dosya kopyalanamadı: {e}")
    exit(1)

# === SİNYALİZASYONU BAŞLAT ===
signal_script = r"C:\Users\Muhammet Samkara\Desktop\Borsa Kripto\signal_generator.py"
if not os.path.exists(signal_script):
    print(f"[ERR] signal_generator.py bulunamadı.")
    exit(1)

try:
    print(f"[6] Sinyalizasyon başlatılıyor...")
    subprocess.run(["python", signal_script, str(target_date)], check=True)
except subprocess.CalledProcessError as e:
    print(f"[ERR] signal_generator.py çalışırken hata oluştu: {e}")
    exit(1)

print(f"[DONE] Tüm işlem başarıyla tamamlandı.")
input("Press any key to continue . . .")
