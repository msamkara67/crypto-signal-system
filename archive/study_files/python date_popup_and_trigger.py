import tkinter as tk
from tkinter import simpledialog, messagebox
from datetime import datetime
import sys
import os

# Geçici dosya yolu
temp_path = "date_input_result.txt"

root = tk.Tk()
root.withdraw()

# Kullanıcıdan tarih al
date_str = simpledialog.askstring("Tarih Girişi", "Lütfen veri çekmek istediğiniz tarihi girin (yyyy-aa-gg):")

# Cancel'a basıldıysa dosyaya yaz ve çık
if date_str is None:
    with open(temp_path, "w") as f:
        f.write("CANCEL")
    sys.exit(0)

# Geçerli tarih mi kontrol et
try:
    target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
except ValueError:
    messagebox.showerror("Hata", "Tarih formatı hatalı. Örnek: 2025-07-17")
    with open(temp_path, "w") as f:
        f.write("INVALID")
    sys.exit(0)

# Excel'e yaz
from openpyxl import load_workbook

excel_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
wb = load_workbook(excel_path)
ws = wb["Daily Update"]
ws.cell(row=1, column=2).value = target_date
wb.save(excel_path)

with open(temp_path, "w") as f:
    f.write("OK")

messagebox.showinfo("Başarılı", f"Tarih {target_date} olarak ayarlandı.")

