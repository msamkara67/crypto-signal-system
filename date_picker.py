import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
from openpyxl import load_workbook
from datetime import datetime
import sys

# Excel yolları
template_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_template.xlsx"
template_sheet = "Daily Update"

rsi_path = r"C:\Users\Muhammet Samkara\Desktop\coin_Updater\coin_data_180days_top100.xlsx"
rsi_sheet = "RSI"

def select_date_and_write():
    def on_select():
        date_str = cal.get_date()
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

            # 1. coin_data_template.xlsx → B1 hücresine
            wb_template = load_workbook(template_path)
            ws_template = wb_template[template_sheet]
            ws_template["B1"].value = date_obj
            wb_template.save(template_path)
            wb_template.close()

            # 2. coin_data_180days_top100.xlsx → RSI sayfası N2 hücresi
            wb_rsi = load_workbook(rsi_path)
            if rsi_sheet in wb_rsi.sheetnames:
                ws_rsi = wb_rsi[rsi_sheet]
                ws_rsi["N2"].value = date_obj
                wb_rsi.save(rsi_path)
                wb_rsi.close()
            else:
                messagebox.showerror("Hata", f"'{rsi_sheet}' sayfası bulunamadı.")

            print(f"{date_obj}")  # stdout'a yaz, ana kod bunu okuyacak
            root.destroy()
        except Exception as e:
            messagebox.showerror("Hata", f"Tarih işlenemedi:\n{e}")

    root = tk.Tk()
    root.title("Tarih Seç")
    root.geometry("300x250")
    root.resizable(False, False)

    cal = Calendar(root, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack(padx=10, pady=10)

    btn = tk.Button(root, text="Tamam", command=on_select)
    btn.pack(pady=10)

    root.mainloop()

# Ana fonksiyonu çalıştır
if __name__ == "__main__":
    select_date_and_write()




