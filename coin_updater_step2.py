
import os
from config import working_dir_name


desktop = os.path.expanduser("~/Desktop")
base_dir = os.path.join(desktop, working_dir_name)
from openpyxl import load_workbook
from datetime import datetime, timedelta
from datetime import datetime as dt, time as dttime
import yfinance as yf
import os
import shutil
import subprocess

# === Pop-up ile tarih al ===
result = subprocess.run(['python', 'date_picker.py'], capture_output=True, text=True)
target_date_str = result.stdout.strip()

if not target_date_str:
    raise ValueError("🛑 Geçerli bir tarih girilmedi.")

target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
print(f"🎯 Veri çekilecek tarih: {target_date}")

# === Dosya Yolları ===

from config import template_path, data_path
sheet_name = "Daily Update"

print("=== Günlük veri çekiliyor ve aktarılıyor ===")
wb_template = load_workbook(template_path)
ws_template = wb_template[sheet_name]

# === Coin Listesini Al ===
coins = []
row_index = 2
while True:
    coin = ws_template.cell(row=row_index, column=3).value
    if coin is None or str(coin).strip() == "":
        break
    coins.append((row_index, coin))
    row_index += 1

# === Tarih Aralığını Ayarla ===
start_datetime = datetime.combine(target_date, dttime.min)
end_datetime = datetime.combine(target_date + timedelta(days=1), dttime.min)

# === Verileri Çek ve Yaz ===
for row_index, coin in coins:
    try:
        ticker = yf.Ticker(f"{coin}-USD")
        hist = ticker.history(start=start_datetime, end=end_datetime)

        if not hist.empty:
            close_val = "{:.8f}".format(hist["Close"].iloc[0])
            vol_val = "{:.2f}".format(hist["Volume"].iloc[0])
            high_val = "{:.8f}".format(hist["High"].iloc[0])
            low_val = "{:.8f}".format(hist["Low"].iloc[0])
        else:
            close_val = "NaN"
            vol_val = "NaN"
            high_val = "NaN"
            low_val = "NaN"
            print(f"⚠ {coin}-USD → Veri bulunamadı.")

    except Exception as e:
        close_val = vol_val = high_val = low_val = "HATA"
        print(f"❌ {coin}-USD → Hata: {e}")

    ws_template.cell(row=row_index, column=4).value = close_val
    ws_template.cell(row=row_index, column=5).value = vol_val
    ws_template.cell(row=row_index, column=6).value = high_val
    ws_template.cell(row=row_index, column=7).value = low_val

wb_template.save(template_path)
print("✔ Template dosyası güncellendi.")

# === 180 Günlük Dosyayı Güncelle ===
wb_data = load_workbook(data_path)
ws_close = wb_data["Daily Update (Close)"]
ws_volume = wb_data["Daily Update (Volume)"]
ws_high = wb_data["Daily Update (High)"]
ws_low = wb_data["Daily Update (Low)"]

def find_column_by_date(ws, date_val):
    for col in range(4, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, datetime) and val.date() == date_val:
            return col
    return None

target_col = find_column_by_date(ws_close, target_date)
if not target_col:
    raise ValueError(f"🛑 Hedef tarih bulunamadı: {target_date}")

for row_index, coin in coins:
    close_val = ws_template.cell(row=row_index, column=4).value
    vol_val = ws_template.cell(row=row_index, column=5).value
    high_val = ws_template.cell(row=row_index, column=6).value
    low_val = ws_template.cell(row=row_index, column=7).value

    for r in range(2, ws_close.max_row + 1):
        if ws_close.cell(row=r, column=3).value == coin:
            ws_close.cell(row=r, column=target_col).value = close_val
            ws_volume.cell(row=r, column=target_col).value = vol_val
            ws_high.cell(row=r, column=target_col).value = high_val
            ws_low.cell(row=r, column=target_col).value = low_val
            break

wb_data.save(data_path)
print("✅ Veriler ana 180 günlük dosyaya aktarıldı.")

# --- XLSM ve Diğer Scriptler ---
print("📦 XLSM versiyonu oluşturuluyor...")
try:
    subprocess.run(["python", "update_high_low.py"], check=True)
    subprocess.run(["python", "XLSM_version.py"], check=True)
    print("✅ XLSM dosyası başarıyla oluşturuldu ve makro uygulandı.")
except subprocess.CalledProcessError as e:
    print(f"❌ XLSM scripti çalıştırılırken hata oluştu: {e}")




